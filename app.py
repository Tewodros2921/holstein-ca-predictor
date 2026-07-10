import streamlit as st
import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
import statsmodels.formula.api as smf
import statsmodels.api as sm
from scipy import stats

# ---------------------------------------------------------
# PAGE CONFIG
# ---------------------------------------------------------
st.set_page_config(
    page_title="Holstein Ca Predictor",
    page_icon="🥛",
    layout="centered"
)

# ---------------------------------------------------------
# TITLE & INTRO
# ---------------------------------------------------------
st.title("🥛 Holstein Calcium Absorption Predictor")
st.markdown("### 🎓 M.Sc. Research Project | The Hebrew University of Jerusalem (HUJI)")
st.info("🔬 Developed at the **Agricultural Research Organization - Volcani Institute**, Department of Ruminant Science.")

st.markdown("""
**Project Objective:**  
This interactive multi-level meta-regression tool calculates the non-linear apparent calcium absorption 
efficiency in lactating Holstein dairy cattle. It translates core metabolic data into actionable physiological 
insights to optimize livestock mineral management and reduce environmental excretion pathways.
""")

# ---------------------------------------------------------
# SIDEBAR INPUTS
# ---------------------------------------------------------
st.sidebar.header("🔬 Input Herd & Feed Metrics")
days_in_milk = st.sidebar.slider("Days in Milk (DIM)", 1, 300, 100, 5)

# ---------------------------------------------------------
# MAIN MODEL INPUT
# ---------------------------------------------------------
st.markdown("---")
st.subheader("🐄 Holstein Cow Calcium Apparent Absorption Model")

ca_intake = st.number_input("Calcium Intake in Feed (X, grams/day):", min_value=1.0, value=100.0, step=5.0)

# ---------------------------------------------------------
# QUADRATIC MODEL
# ---------------------------------------------------------
intercept = 62.47
beta_x = -0.2611
beta_x_sq = 0.000561
beta_dim = -0.0202

predicted_y = (
    intercept +
    (beta_x * ca_intake) +
    (beta_x_sq * (ca_intake ** 2)) +
    (beta_dim * days_in_milk)
)

predicted_y = max(0.0, min(100.0, predicted_y))

# ---------------------------------------------------------
# MASS BALANCE
# ---------------------------------------------------------
calculated_absorbed_g = (predicted_y / 100.0) * ca_intake
calculated_feces = ca_intake - calculated_absorbed_g

# ---------------------------------------------------------
# KPI METRICS
# ---------------------------------------------------------
col_res1, col_res2, col_res3 = st.columns(3)
with col_res1:
    st.metric("📥 Intake Calcium (X)", f"{ca_intake:.1f} g")
with col_res2:
    st.metric("💩 Calculated Fecal Calcium", f"{calculated_feces:.1f} g", delta="Excreted Waste", delta_color="inverse")
with col_res3:
    st.metric("✅ Predicted Absorbed Mass", f"{calculated_absorbed_g:.1f} g", delta="Retained Balance")

# ---------------------------------------------------------
# MODEL RESULT
# ---------------------------------------------------------
st.markdown("---")
st.subheader("🔮 Model Prediction Result (Y)")
st.metric("Predicted Apparent Calcium Absorption Efficiency (Y)", f"{predicted_y:.2f} %")

st.markdown("### 🧮 Quadratic Model Regression Formula")
st.latex(r"Y = 62.47 + (-0.2611 \times X) + (0.000561 \times X^2) + (-0.0202 \times \text{DIM})")

st.info(f"**Live Equation Process:** 62.47 + (-0.2611 × {ca_intake}) + (0.000561 × {ca_intake}²) + (-0.0202 × {days_in_milk}) = **{predicted_y:.2f}%**")
st.success(f"**Fecal Mass Derivation:** {ca_intake}g Intake - ({predicted_y:.2f}% Absorption × {ca_intake}g) = **{calculated_feces:.1f} grams of Fecal Calcium**")

# ---------------------------------------------------------
# CLEAN IF / ELIF / ELSE BLOCK
# ---------------------------------------------------------
if predicted_y < 25.0:
    st.warning("⚠️ Low Predicted Efficiency: High risk of mineral pass-through and environmental fecal excretion.")
elif predicted_y > 55.0:
    st.info("📈 High Predicted Activity: Intestinal active transport channels upregulated (high metabolic draw).")
else:
    st.success("✅ Normal Range: Physiological baseline absorption levels maintained.")

# ---------------------------------------------------------
# OPTIONAL DATASET CALCULATION BLOCK
# ---------------------------------------------------------
if "df" in globals():
    if "Ca_Intake" in df.columns and "Fecal_Ca" in df.columns:
        df["Apparent_Ca_Absorption"] = (
            (df["Ca_Intake"] - df["Fecal_Ca"]) / df["Ca_Intake"]
        )
        st.success("Apparent Calcium Absorption calculated successfully.")
    else:
        st.warning("Dataset must include Ca_Intake and Fecal_Ca columns.")

# ---------------------------------------------------------
# ADVANCED EXCEL‑STYLE SCHEDULING SECTION
# ---------------------------------------------------------
st.markdown("---")
st.header("📅 Advanced Excel Scheduling Form")

schedule = pd.DataFrame({
    "Time": [
        "06:00", "07:00", "08:00", "09:00", "10:00",
        "11:00", "12:00", "13:00", "14:00", "15:00",
        "16:00", "17:00", "18:00"
    ],
    "Task": [""] * 13
})

edited_schedule = st.data_editor(schedule, num_rows="dynamic")

if st.button("Download Schedule (.xlsx)"):
    import openpyxl
    from openpyxl.styles import Alignment, Font

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Daily Schedule"

    ws["A1"] = "Time"
    ws["B1"] = "Task"
    ws["A1"].font = Font(bold=True)
    ws["B1"].font = Font(bold=True)

    for i, row in edited_schedule.iterrows():
        ws[f"A{i+2}"] = row["Time"]
        ws[f"B{i+2}"] = row["Task"]
        ws[f"A{i+2}"].alignment = Alignment(horizontal="center")
        ws[f"B{i+2}"].alignment = Alignment(horizontal="left")

    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 40
    # ---------------------------------------------------------
# ALARM SOUND (UPLOAD YOUR OWN RECORDING)
# ---------------------------------------------------------
st.markdown("---")
st.header("🔊 Custom Alarm Sound")

uploaded_sound = st.file_uploader(
    "Upload your alarm sound (MP3 or WAV):",
    type=["mp3", "wav"]
)

if uploaded_sound is not None:
    import base64

    sound_bytes = uploaded_sound.read()
    b64 = base64.b64encode(sound_bytes).decode()

    st.success("Alarm sound uploaded successfully!")

    if st.button("Play Alarm"):
        audio_html = f"""
            <audio autoplay>
                <source src="data:audio/mp3;base64,{b64}" type="audio/mp3">
            </audio>
        """
        st.markdown(audio_html, unsafe_allow_html=True)


    wb.save("schedule.xlsx")
    st.success("Saved schedule.xlsx")
